import openpyxl


class HomePageData :


    def getHomePageData (test_case_name):

        book= openpyxl.load_workbook("C:\\Users\\raa918070\\Desktop\\DataDrivenFM\\Eguru_HomePage.xlsx")
        sheet= book.active
        for i in range(1, sheet.max_row + 1) :  # to get rows
            if sheet.cell(row=i, column=1).value == test_case_name :

                for j in range(2, sheet.max_column + 1) :  # to get columns
                    # Dict["lastname"]="shetty
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return [Dict]
